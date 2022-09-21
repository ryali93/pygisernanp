# -*- coding: utf-8 -*-
import arcpy
import os
import numpy as np
import pandas as pd
from textwrap import wrap
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl import drawing
import os
import string
import matplotlib.pyplot as plt

plt.style.use('ggplot')
plt.style.use('seaborn-white')

arcpy.env.overwriteOutput = True

fuente_dict = {
    1: "Planet",
    2: "GLADS2",
    3: "PNCBLandsat"
}

def abc(inicial, final):
    abc = list(string.ascii_uppercase)
    ini = abc.index(inicial)
    fin = abc.index(final)
    lista = abc[ini:fin + 1]
    return lista

def get_data(gdb, fc, fuente, mesrep):
    fields = ["anp_codi", "zi_codi", "md_sup", "md_fecimg", "md_bosque", "md_causa"]
    domains = [x for x in arcpy.da.ListDomains(gdb) if x.name == "ANP"][0]
    sql = "md_mesrep = {} AND md_fuente = {}".format(mesrep, fuente)
    # sql = "md_mesrep IN (2022034, 2022042)  AND md_fuente = 3"
    np_df = arcpy.da.TableToNumPyArray(fc, fields, sql)
    df = pd.DataFrame(np_df)
    df["anp_nomb"] = df["anp_codi"].apply(lambda x: domains.codedValues[x])
    
    df2 = df[df["md_causa"] != 14]

    df_f = df2.groupby(("anp_codi", "zi_codi")).agg({
        "md_sup": "sum",
        "md_fecimg": "first",
        "anp_nomb": "first"
    })
    df_n = df_f.reset_index()
    data_1 = pd.DataFrame({
        "anp_nomb": df_n["anp_nomb"],
        "anp_codi": df_n["anp_codi"],
        "zi_codi": df_n["zi_codi"],
        "md_fecimg": df_n["md_fecimg"].apply(lambda x: x.strftime("%d-%m-%Y")),
        "md_sup": df_n["md_sup"].apply(lambda x:round(x,2))
    })
    data_1 = data_1[["anp_nomb", "anp_codi", "zi_codi", "md_fecimg", "md_sup"]]
    data_2 = df2.groupby("anp_nomb")["md_sup"].sum()
    data_2 = data_2.reset_index()
    data_2["md_sup"] = data_2["md_sup"].apply(lambda x:round(x,2))

    total = ["Total general", round(data_2["md_sup"].sum(), 2)]

    data_1 = data_1.values.tolist()
    data_2 = data_2.values.tolist()
    data_2.append(total)

    msg_dates = "{} de {} al {} de {}".format(
        min(df_n["md_fecimg"]).strftime("%d"),
        min(df_n["md_fecimg"]).strftime("%B"),
        max(df_n["md_fecimg"]).strftime("%d"),
        max(df_n["md_fecimg"]).strftime("%B")
    )

    ## Bosque
    df_bosque = df.groupby("md_bosque")["md_sup"].sum()
    if 2 in df_bosque.keys():
        data_bosque = [round(df_bosque[2], 2), round(df_bosque[2]/df_bosque.sum()*100, 2)]
    else:
        data_bosque = []

    ## Causa
    df_causa = df.groupby("md_causa")["md_sup"].sum()
    if 14 in df_causa.keys():
        data_causa = [round(df_causa[14], 2), round(df_causa[14] / df_causa.sum() * 100, 2)]
    else:
        data_causa = []

    return data_1, data_2, msg_dates, data_bosque, data_causa


def create_graph(df, msg_dates, msg_title, path_graph):
    '''
    :param df: DataFrame con info de lo necesario para realizar el grafico resumen
    :return:
    '''
    df_n = pd.DataFrame(df)
    df_n = df_n[:-1]
    df_n = df_n.sort_values(by=[1], ascending=False)
    large_letter = df_n[0].tolist()
    large_num = df_n[1].tolist()
    width = 1.0
    lefts = [x * width for x, _ in enumerate(large_num)]
    large_letter = ['\n'.join(wrap(x, 15)) for x in large_letter]

    msg_title_n = '\n'.join(wrap(msg_title, 70))
    plt.rcdefaults()
    plt.style.use('ggplot')
    plt.rcParams.update({'figure.autolayout': True})
    fig, ax = plt.subplots(figsize=(8, 10))
    ax.barh(lefts, large_num, tick_label=large_letter, align='center')
    ax.set_title(msg_title_n, fontdict={'fontsize': 11, 'fontweight': 'medium'})
    plt.xlabel(u'Suma de hectáreas', fontsize=10)
    labels = ax.get_xticklabels()
    plt.savefig(path_graph)

def create_msg_title(fuente, msg_dates):
    msg_fuente = ""
    if fuente == 1: # Planet
        msg_fuente = u'Monitoreo con Imágenes Planet'
    elif fuente == 2: # GLAD S2
        msg_fuente = u'Alertas GLAD-S2 de la Universidad de Maryland'
    elif fuente == 3: # PNCB
        msg_fuente = u'ATD del PNCB con filtro de la causa de pérdida'

    msg_title = u'Deforestación en hectáreas en ANP del bioma amazónico del {} - {} ' \
                 u'realizado por el área de teledetección de SERNANP'.format(msg_dates,
                                                                             msg_fuente)
    return msg_title

def create_report(df1, df2, path_xlsx_template, path_graph, path_xlsx_outh, msg_dates, msg_title, data_bosque, data_causa, fuente):
    wb = load_workbook(path_xlsx_template)
    ws = wb.get_sheet_by_name("Reporte")

    ws["A4"].value = msg_title
    ws.insert_rows(7, len(df1))

    thin_border = Border(bottom=Side(style='thin'))

    x = 0
    for x in range(7, 7+len(df1)):
        ws["A{}".format(x)] = df1[x-7][0]
        ws["B{}".format(x)] = df1[x-7][1]
        ws["C{}".format(x)] = df1[x-7][2]
        ws["D{}".format(x)] = df1[x-7][3]
        ws["E{}".format(x)] = df1[x-7][4]
        ws["A{}".format(x)].alignment = Alignment(vertical='center')
        ws["B{}".format(x)].alignment = Alignment(horizontal='center', vertical='center')
        ws["C{}".format(x)].alignment = Alignment(horizontal='center', vertical='center')
        ws["D{}".format(x)].alignment = Alignment(horizontal='center', vertical='center')
        ws["E{}".format(x)].alignment = Alignment(horizontal='center', vertical='center')
        ws["A{}".format(x)].border = thin_border
        ws["B{}".format(x)].border = thin_border
        ws["C{}".format(x)].border = thin_border
        ws["D{}".format(x)].border = thin_border
        ws["E{}".format(x)].border = thin_border

    # FUENTE
    nrow = 7 + len(df1) + 2
    if fuente == 1:
        ws["A{}".format(nrow)] = ""
        ws["A{}".format(nrow+1)] = ""
    elif fuente == 2:
        ws["A{}".format(nrow)] = u"1) Pérdida forestal en hectáreas en ANPs {} del 2022 - Alertas GLAD S2".format(msg_dates)
    elif fuente == 3:
        ws["A{}".format(nrow)] = u"1) Pérdida forestal en hectáreas en ANPs {} del 2022 - ATDs del PNCB".format(msg_dates)

    # BOSQUE
    nrow = 7 + len(df1) + 8
    if len(data_bosque) > 1:
        texto = "5) "
        ws["A{}".format(nrow)] = texto + "{} ha ({}% de las alertas de este reporte) corresponden " \
                                         "a deforestación en bosques secundarios".format(data_bosque[0],
                                                                                         data_bosque[1])
        nrow = nrow + 1
    else:
        texto = ""
        ws["A{}".format(nrow)] = texto

    # CAUSA
    if len(data_causa) > 1:
        texto = "6) " if texto == "" else "5) "
        ws["A{}".format(nrow)] = texto + "{} ha ({}% de las alertas) corresponden a pérdidas boscosas " \
                                 "por causa natural".format(data_causa[0],
                                                            data_causa[1])
    else:
        texto = ""
        ws["A{}".format(nrow)] = texto

    nrow = 7 + len(df1) + 12

    for x in range(nrow, nrow+len(df2)):
        ws["A{}".format(x)] = df2[x-nrow][0]
        ws["B{}".format(x)] = df2[x-nrow][1]
        ws["B{}".format(x)].alignment = Alignment(horizontal='center', vertical='center')
        ws["A{}".format(x)].border = thin_border
        ws["B{}".format(x)].border = thin_border
    ws["A{}".format(x)].font = Font(bold=True)
    ws["B{}".format(x)].font = Font(bold=True)

    nrow_n = nrow + len(df2) + 2

    img = drawing.image.Image(path_graph)
    img.height = 300
    img.width = 420
    ws.add_image(img, 'A{}'.format(nrow_n))

    wb.save(path_xlsx_outh)

def proccess():
    gdb = r'Database Connections\gdb.sde'
    fc = os.path.join(gdb, r'gdb.sde.MonitoreoCobertura\gdb.sde.MonitoreoDeforestacion') # os.path.join(gdb, r'MonitDefor')
    path_xlsx_template = r'F:\sernanp\proyectos\monitoreo\reporte\reporte.xlsx'
    path_xlsx_outh = r'F:\sernanp\proyectos\monitoreo\reporte\reporte_2022065_pncb.xlsx'

    temp_folder = tempfile.mkdtemp()
    path_graph = os.path.join(temp_folder, "reporte_tmp.png")

    fuente = 3
    mesrep = 2022065
    df1, df2, msg_dates, data_bosque, data_causa = get_data(gdb, fc, fuente=fuente, mesrep=mesrep)
    msg_title = create_msg_title(fuente, msg_dates)
    create_graph(df2, msg_dates, msg_title, path_graph)
    create_report(df1, df2, path_xlsx_template, path_graph, path_xlsx_outh, msg_dates, msg_title, data_bosque, data_causa, fuente)

def main():
    proccess()

if __name__ == '__main__':
    main()
